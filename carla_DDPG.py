import glob
import os
import sys
import random
import time
import numpy as np
import math
import torch
import nn_actic_critic2
# import environment as env
import suiji as env
import utility as ut
import model1 as md
from keras.models import load_model
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import datetime
SHOW_PREVIEW = True
np.random.seed(32)
random.seed(32)
AGGREGATE_STATS_EVERY = 4
ROUNDING_FACTOR = 2
wb = load_workbook('demo6.xlsx')
sheet = wb.active

sheet['a1'] = 'actor_loss'
sheet['b1'] = 'critic_loss'
sheet['c1'] = 'avg_reward'
sheet['d1'] = 'max_reward'
sheet['e1'] = 'min_reward'


try:
	sys.path.append(glob.glob('../../../carla/dist/carla-*%d.%d-%s.egg' % (
		sys.version_info.major,
		sys.version_info.minor,
		'win-amd64' if os.name == 'nt' else 'linux-x86_64'))[0])
except IndexError:
	pass


import carla


def choose_action_lane_change(choice):
	action = []
	action = [0.5, float(choice), 0, False]
	return action

def train_left_lane_change_DDPG(episodes,agent):
	action_space = 1
	state_space = 5
	radar_space = 500
	leftLaneChange_model = md.DDPG(action_space, state_space, radar_space, 'Lane_Change_Model')
	tau = 0.005
	ep_reward_list = []
	avg_reward_list = []
	max_reward_list = []
	min_reward_list = []
	score_list = []
	actor_loss = []
	critic_loss = []
	for epi in range(episodes):
		try:
			loc = random.randint(30, 130)
			print(f'--------Spawn Succeded Lane Change-----------')
			radar_state_prev = agent.reset(True)
			radar_state_prev = np.reshape(radar_state_prev, (1, radar_space))
			start_state = [0, 0, 0, 0, 19]
			state = np.reshape(start_state, (1, state_space))
			score = 0
			max_step = 5_00
			actor_loss_epi = []
			critic_loss_epi = []
			for i in range(max_step):
				choice = leftLaneChange_model.policy(radar_state_prev, state)
				action = choose_action_lane_change(choice)
				print(f'action----{action}-------epsilon----{leftLaneChange_model.epsilon}')
				radar_state_next, next_state, reward, done, _ = agent.left_lane_change(action)
				time.sleep(0.2)
				score += reward
				next_state = np.reshape(next_state, (1, state_space))
				leftLaneChange_model.remember(radar_state_prev, radar_state_next, state, choice, reward, next_state, done)
				state = next_state
				radar_state_prev = np.reshape(radar_state_next, (1, radar_space))
				lossActor, lossCritic = leftLaneChange_model.replay()
				actor_loss_epi.append(lossActor)
				critic_loss_epi.append(lossCritic)
				leftLaneChange_model.update_target(tau, 1)
				if done:
					break

			actor_loss.append(np.mean(actor_loss_epi))
			critic_loss.append(np.mean(critic_loss_epi))
			leftLaneChange_model.update_target(0.001, epi)
			ep_reward_list.append(score)
			# ep_reward_list.append(reward_1)
			print("\nepisode: {}/{}, score: {}".format(epi, episodes, score))
			avg_reward = np.mean(ep_reward_list[-AGGREGATE_STATS_EVERY:])
			print("\nEpisode * {} * Avg Reward is ==> {}\n".format(epi, avg_reward))
			avg_reward_list.append(avg_reward)
			min_reward = min(ep_reward_list[-AGGREGATE_STATS_EVERY:])
			max_reward = max(ep_reward_list[-AGGREGATE_STATS_EVERY:])
			max_reward_list.append(max_reward)
			min_reward_list.append(min_reward)
			score_list.append(score)
			#  leftLaneChange_model.tensorboard.update_stats(reward_avg=avg_reward, critic_loss=np.mean(critic_loss_epi), actor_loss=np.mean(actor_loss_epi))
			leftLaneChange_model.tensorboard.update_stats(reward_avg= avg_reward, critic_loss=np.mean(critic_loss_epi), actor_loss=np.mean(actor_loss_epi))

			if(epi%500==0 and epi>1):
				x_label = 'Episodes'
				y_label = 'Actor Loss'
				ut.plot(actor_loss, x_label, y_label, epi)
				y_label = 'Critic Loss'
				ut.plot(critic_loss,  x_label, y_label, epi)
				time.sleep(1)

		finally:
			print(f"Task Completed! Episode {epi}")
			leftLaneChange_model.save_model()
			if agent != None:
				agent.destroy()
				time.sleep(2)

	return actor_loss, critic_loss, avg_reward_list, max_reward_list, min_reward_list, score_list


if __name__ == '__main__':
	agent = env.CarlaVehicle()
	episodes = 5000
	actor_Loss, critic_Loss, avg_reward, max_reward, min_reward, scores = train_left_lane_change_DDPG(episodes, agent)
	for i in range(episodes):
		sheet.cell(row=i+1, column=1).value = actor_Loss[i]
		sheet.cell(row=i+1, column=2).value = critic_Loss[i]
		sheet.cell(row=i+1, column=3).value = avg_reward[i]
		sheet.cell(row=i+1, column=4).value = max_reward[i]
		sheet.cell(row=i+1, column=5).value = min_reward[i]
		sheet.cell(row=i+1, column=6).value = scores[i]
		sheet.cell(row=i+1, column=7).value = i
		wb.save('demo8.xlsx')
	print("\n\n--We need to Maxmise Actor Loss--Minimise Critic Loss--\n\n")
	x_label = 'Episodes'
	y_label = 'Actor Loss'
	ut.plot(actor_Loss, x_label, y_label, epi=5000)
	plt.show()
	y_label = 'Critic Loss' 
	ut.plot(critic_Loss,  x_label, y_label, epi=5000)
	plt.show()
	y_label = 'Reward'
	ut.pllot(avg_reward, x_label, y_label, epi=5000)
	plt.show()
	y_label = 'Scores'
	ut.pllot(scores, x_label, y_label, epi=5000)
	plt.show()
